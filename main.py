# -*- coding: utf-8 -*-
from parse import Parse
import requests
import time
import logging
import codecs
import openpyxl

logging.basicConfig(level=logging.ERROR,
                    format='%(asctime)s Process%(process)d:%(thread)d %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    filename='diary.log',
                    filemode='a')


def init():
    headers = {
        'Connection': 'keep-alive',
        'Cache-Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    }
    response = requests.get(
        'https://www.lagou.com/jobs/list_?city=%E4%B8%8A%E6%B5%B7&cl=false&fromSearch=true&labelWords=&suginput=',
        headers=headers)  # 请求原网页
    r = requests.utils.dict_from_cookiejar(response.cookies)  # 获取cookies
    print(r)
    r["user_trace_token"] = r["LGRID"]
    r["LGSID"] = r["LGRID"]
    r["LGUID"] = r["LGRID"]  # 构造cookies的参数
    cookies = {
        'X_MIDDLE_TOKEN': '797bc148d133274a162ba797a6875817',
        'JSESSIONID': 'ABAAABAAAIAACBI03F33A375F98E05C5108D4D742A34114',
        '_ga': 'GA1.2.1912257997.1548059451',
        '_gat': '1',
        'Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6': '1548059451',
        'user_trace_token': '20190121163050-dbd72da2-1d56-11e9-8927-525400f775ce',
        'LGSID': '20190121163050-dbd72f67-1d56-11e9-8927-525400f775ce',
        'PRE_UTM': '',
        'PRE_HOST': '',
        'PRE_SITE': '',
        'PRE_LAND': 'https%3A%2F%2Fwww.lagou.com%2F%3F_from_mid%3D1',
        'LGUID': '20190121163050-dbd73128-1d56-11e9-8927-525400f775ce',
        '_gid': 'GA1.2.1194828713.1548059451',
        'index_location_city': '%E5%85%A8%E5%9B%BD',
        'TG-TRACK-CODE': 'index_hotjob',
        'LGRID': '20190121163142-fb0cc9c0-1d56-11e9-8928-525400f775ce',
        'Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6': '1548059503',
        'SEARCH_ID': '86ed37f5d8da417dafb53aa25cd6fbc0',
    }
    cookies.update(r)  # 更新接口的cookies
    headers = {
        'Origin': 'https://www.lagou.com',
        'X-Anit-Forge-Code': '0',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Referer': 'https://www.lagou.com/jobs/list_java?px=new&city=%E4%B8%8A%E6%B5%B7',
        'X-Requested-With': 'XMLHttpRequest',
        'Connection': 'keep-alive',
        'X-Anit-Forge-Token': 'None',
    }
    params = (
        ('px', 'new'),
        ('city', '\u4E0A\u6D77'),
        ('needAddtionalResult', 'false'),
    )
    data = {'first': True,
            'kd': 'java',
            'pn': 1}

    res = {}
    res['data'] = data
    res['cookies'] = cookies
    res['headers'] = headers
    res['params'] = params
    return res

def getInfo(url, para):
    """
    获取信息
    """

    res = init()
    res = requests.post('https://www.lagou.com/jobs/positionAjax.json', headers=res['headers'], params=res['params'],
                             cookies=res['cookies'], data=res['data'])  # 请求接口
    htmlCode = res.text
    generalParse = Parse(htmlCode)
    pageCount = generalParse.parsePage()
    print('pageCount',pageCount)
    info = []
    for i in range(1, pageCount + 1):
    # for i in range(1,2):
        print('第%s页' % i)
        para['pn'] = str(i)
        res = init()
        res = requests.post('https://www.lagou.com/jobs/positionAjax.json', headers=res['headers'], params=res['params'],cookies=res['cookies'], data=res['data'])  # 请求接口
        htmlCode = res.text
        generalParse = Parse(htmlCode)
        info = info + getInfoDetail(generalParse)
        time.sleep(3)
    return info


def getInfoDetail(generalParse):
    """
    信息解析
    """
    info = generalParse.parseInfo()
    return info


def processInfo(info, para):
    """
    信息存储
    """
    logging.error('Process start')
    try:
        title = '公司名称\t公司类型\t融资阶段\t标签\t公司规模\t公司所在地\t职位类型\t学历要求\t福利\t薪资\t工作经验\n'
        # file = codecs.open('%s职位.xls' % para['city'], 'w', 'utf-8')
        wb = openpyxl.Workbook('python1.xlsx')
        sht2 = wb.create_sheet(title='s2')
        rownum = 1
        for p in info:
            columnnum = 1
            for key in p.keys():
                if isinstance(p[key],list):
                    res = ''.join(p[key])
                    sht2.cell(row=rownum, column=columnnum,value=res)
                else:
                    sht2.cell(row=rownum, column=columnnum, value=p[key])
                columnnum += 1
            rownum += 1
            wb.save('python1.xlsx')
            wb.close()
        return True
    except Exception as e:
        print(e)
        return None


def main(url, para):
    """
    主函数逻辑
    """
    logging.error('Main start')
    if url:
        info = getInfo(url, para)  # 获取信息
        flag = processInfo(info, para)  # 信息储存
        return flag
    else:
        return None


if __name__ == '__main__':
    kdList = [u'python']
    cityList = [u'上海']
    url = 'https://www.lagou.com/jobs/positionAjax.json?city=上海&needAddtionalResult=false'
    for city in cityList:
        print('爬取%s' % city)
        para = {'first': 'true', 'pn': '1', 'kd': kdList[0], 'city': city}
        flag = main(url, para)
        if flag:
            print('%s爬取成功' % city)
        else:
            print('%s爬取失败' % city)
